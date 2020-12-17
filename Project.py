import pymongo
import openpyxl
client=pymongo.MongoClient()
Students=client.Student_Database
Teachers=client.Teacher_Database
Rooms=client.Room_Database
Rooms_Students=client.Database
Courses=client.Course_Database
Teacher_room=client.Teacher_duty_Database
Teacher_room_table=Teacher_room.Table
student_table=Students.Table
course_table=Courses.Table
Teacher_table=Teachers.Table
Room_table=Rooms.Table
Room_student=Rooms_Students.Table
Type_Entry=client.Type
Type_thing=Type_Entry.Table
Document=openpyxl.load_workbook("Exam allotment.xlsx")
student_details=Document['Sheet1']
teacher_details=Document['Sheet2']
room_details=Document['Sheet3']
class Input():
    global Dates
    global Time_1
    global Times
    Dates=[]
    Times=[]        
    Time_1=None
    Type=None
    def Type_Dates(self):
        Type=int(input("Enter 1 for Internals\nEnter 2 for End Semester Exams\n\n"))
        if Type==1:
            print("Enter 4 dates for Exams and Two Timings for each Date\n")
            print("Enter Dates is DD/MM/YY format")
            for date in range(0,4):
                Dates.append(str(input("Enter the Date")))
            for date in range(0,2):
                Times.append(str(input("Enter the first date\n(Enter in HH:MM format)")))            
        else:
            print("Enter 7 dates for Exams and a Time for the Exams\n")
            print("Enter Dates is DD/MM/YY format")
            for date in range(0,7):
                Dates.append(str(input("Enter the Date")))
            Time_1=str(input("Enter the first date\n(Enter in HH:MM format)"))
            dictionary={'Type':Time_1}
            Type_thing.insert_one(dictionary)
            
class Database():
    def __init__(self):
        student_table.drop()
        Teacher_table.drop()
        Room_table.drop()
        Room_student.drop()
        course_table.drop()
        Teacher_room_table.drop()
        Type_thing.drop()
        __Document=openpyxl.load_workbook("Exam allotment.xlsx")
        __student_details=__Document['Sheet1']
        __teacher_details=__Document['Sheet2']
        __room_details=__Document['Sheet3']
        __BCA_1=__Document['Sheet4']
        __BCA_2=__Document['Sheet5']
        __BCA_3=__Document['Sheet6']
        __BCA_4=__Document['Sheet7']
        __BCA_5=__Document['Sheet8']
        __CSE_1=__Document['Sheet15']
        __CSE_2=__Document['Sheet16']
        __CSE_3=__Document['Sheet9']
        __CSE_4=__Document['Sheet10']
        __CSE_5=__Document['Sheet11']
        __CSE_6=__Document['Sheet12']
        __CSE_7=__Document['Sheet13']
        __CSE_8=__Document['Sheet14']
        for row in range(5,__BCA_1.max_row+1):
            if __BCA_1.cell(row,2).value==None:
                continue
            entry={'Branch':'BCA',"Semester":1,'Course':__BCA_1.cell(row,2).value,'Course_code':__BCA_1.cell(row,1).value,'Date':False,'Time':False}
            course_table.insert_one(entry)
        for row in range(3,__BCA_2.max_row+1):
            if __BCA_2.cell(row,2).value==None:
                continue
            entry={'Branch':'BCA',"Semester":2,'Course':__BCA_2.cell(row,2).value,'Course_code':__BCA_2.cell(row,1).value,'Date':False,'Time':False}
            course_table.insert_one(entry)
        for row in range(3,__BCA_3.max_row+1):
            if __BCA_3.cell(row,2).value==None:
                continue
            entry={'Branch':'BCA',"Semester":3,'Course':__BCA_3.cell(row,2).value,'Course_code':__BCA_3.cell(row,1).value,'Date':False,'Time':False}
            course_table.insert_one(entry)
        for row in range(3,__BCA_4.max_row+1):
            if __BCA_4.cell(row,2).value==None:
                continue
            entry={'Branch':'BCA',"Semester":4,'Course':__BCA_4.cell(row,2).value,'Course_code':__BCA_4.cell(row,1).value,'Date':False,'Time':False}
            course_table.insert_one(entry)
        for row in range(3,__BCA_1.max_row+1):
            if __BCA_5.cell(row,2).value==None:
                continue
            entry={'Branch':'BCA',"Semester":5,'Course':__BCA_5.cell(row,2).value,'Course_code':__BCA_5.cell(row,1).value,'Date':False,'Time':False}
            course_table.insert_one(entry)    
        for row in range(3,__CSE_1.max_row+1):
            if __CSE_1.cell(row,2).value==None:
                continue
            entry={'Branch':'CSE',"Semester":1,'Course':__CSE_1.cell(row,2).value,'Course_code':__CSE_1.cell(row,1).value,'Date':False,'Time':False}
            course_table.insert_one(entry)  
        for row in range(3,__CSE_2.max_row+1):
            if __CSE_2.cell(row,2).value==None:
                continue
            entry={'Branch':'CSE',"Semester":2,'Course':__CSE_2.cell(row,2).value,'Course_code':__CSE_2.cell(row,1).value,'Date':False,'Time':False}
            course_table.insert_one(entry)
        for row in range(3,__CSE_3.max_row+1):
            if __CSE_3.cell(row,2).value==None:
                continue
            entry={'Branch':'CSE',"Semester":3,'Course':__CSE_3.cell(row,2).value,'Course_code':__CSE_3.cell(row,1).value,'Date':False,'Time':False}
            course_table.insert_one(entry)
        for row in range(3,__CSE_4.max_row+1):
            if __CSE_4.cell(row,2).value==None:
                continue
            entry={'Branch':'CSE',"Semester":4,'Course':__CSE_4.cell(row,2).value,'Course_code':__CSE_4.cell(row,1).value,'Date':False,'Time':False}
            course_table.insert_one(entry)    
        for row in range(3,__CSE_5.max_row+1):
            if __CSE_5.cell(row,2).value==None:
                continue
            entry={'Branch':'CSE',"Semester":5,'Course':__CSE_5.cell(row,2).value,'Course_code':__CSE_5.cell(row,1).value,'Date':False,'Time':False}
            course_table.insert_one(entry)
        for row in range(3,__CSE_6.max_row+1):
            if __CSE_6.cell(row,2).value==None:
                continue
            entry={'Branch':'CSE',"Semester":6,'Course':__CSE_6.cell(row,2).value,'Course_code':__CSE_6.cell(row,1).value,'Date':False,'Time':False}
            course_table.insert_one(entry)
        for row in range(3,__CSE_7.max_row+1):
            if __CSE_7.cell(row,2).value==None:
                continue
            entry={'Branch':'CSE',"Semester":7,'Course':__CSE_7.cell(row,2).value,'Course_code':__CSE_7.cell(row,1).value,'Date':False,'Time':False}
            course_table.insert_one(entry)
        for row in range(3,__CSE_8.max_row+1):
            if __CSE_8.cell(row,2).value==None:
                continue
            entry={'Branch':'CSE',"Semester":8,'Course':__CSE_8.cell(row,2).value,'Course_code':__CSE_8.cell(row,1).value,'Date':False,'Time':False}
            course_table.insert_one(entry)
        for row in range(4,__student_details.max_row+1):
            __Student_name=student_details.cell(row,1).value
            __Student_USN=student_details.cell(row,2).value
            __Student_Branch=student_details.cell(row,3).value
            __Student_Semester=student_details.cell(row,4).value
            __Student_row={'Name':__Student_name,'USN':__Student_USN,'Branch':__Student_Branch,'Semester':__Student_Semester,'Alloted':False }
            student_table.insert_one(__Student_row)
        for row in range(5,__teacher_details.max_row + 1):
            __Teacher_name=teacher_details.cell(row,2).value
            __Course_1=teacher_details.cell(row,3).value
            __Course_2=teacher_details.cell(row,4).value
            if((teacher_details.cell(row,5).value)=='Associate professor'):
                __Teacher_Duties=2
            elif((teacher_details.cell(row,5).value)=='Assistant professor'):
                __Teacher_Duties=3
            else:
                __Teacher_Duties=5
            __Teacher_row={'Name':__Teacher_name,'Course 1':__Course_1,'Course 2':__Course_2,'Number of Duties':__Teacher_Duties}
            Teacher_table.insert_one(__Teacher_row)
        for row in range(5,__room_details.max_row+1):
            __Room_Number=room_details.cell(row,1).value
            __Room_Capacity=int(room_details.cell(row,2).value)
            __Room_Location=room_details.cell(row,4).value
            __Room_Type=room_details.cell(row,5).value
            __Room_row={'Room Number':__Room_Number,'Capacity':__Room_Capacity,'Block':__Room_Location,'Type':__Room_Type}
            Room_table.insert_one(__Room_row)            
Example=Database()
class Allocation(Input):
    global Dates
    global Time_1
    global Times
    global Type
    def function(self):
        for room in Room_table.find():
            # print('Room number:\t'+str(room['Room Number'])+'\tBlock:\t'+room['Block'])
            original_room=room.copy()
            original_capacity=room['Capacity']
            branch=None
            semester=None
            while room['Capacity']>0 and bool(student_table.find_one({'Alloted':False})):               
                for student in student_table.find({'Alloted':False},{'_id':False},limit=int(original_capacity/2)).sort('USN'):
                    original_student=student.copy()
                    if room['Capacity']>0:
                        branch=student['Branch']
                        semester=student['Semester']
                        student['Alloted']=True
                        room['Capacity']=room['Capacity']-1
                        Insert={'Room':room['Room Number'],'block':room['Block'],'student':student['USN'],'Branch':student['Branch'],'Semester':student['Semester']}
                        Room_student.insert_one(Insert)
                        changed_student=student.copy()
                        
                        # print('Name:'+student['Name'].title()+'\tUSN:'+student['USN']+'\t:Branch:'+student['Branch']+'\tSemester:'+str(student['Semester']))
                        student_table.find_one_and_replace(original_student,changed_student)
                for students in student_table.find({'Alloted':False,'$or':[{'Branch':{'$ne':branch}},{'Semester':{'$ne':semester}}]},{'_id':False},limit=int(original_capacity/2)):
                    if room['Capacity']>0:
                        original_dict=students.copy()
                        Insert={'Room':room['Room Number'],'block':room['Block'],'student':students['USN'],'Branch':students['Branch'],'Semester':students['Semester']}
                        Room_student.insert_one(Insert)
                        students['Alloted']=True   
                        room['Capacity']=room['Capacity']-1
                        changed_student=students.copy()
                        student_table.find_one_and_replace(original_dict,changed_student)
                    # print('Name:'+students['Name'].title()+'\tUSN:'+students['USN']+'\t:Branch'+students['Branch']+'\tSemester:'+str(students['Semester']))
            changed_room=room.copy()
            Room_table.find_one_and_replace(original_room,changed_room)
        if bool(Type_thing.find_one()):
            iterator=0
            while(course_table.find_one({'Time':False})):
                course=course_table.find_one({'Date':False})
                old=course.copy()
                old_branch=[]
                old_semester=[]
                course['Date']=Dates[iterator]
                course['Time']=Time_1
                old_branch.append(course['Branch'])
                old_semester.append(course['Semester'])
                new=course.copy()
                course_table.find_one_and_replace(old,new)
                while(list(course_table.find({'$or':[{'Branch':{'$nin':old_branch}},{'Semester':{'$nin':old_semester}}],'Date':False}))):
                    course=course_table.find_one({'$or':[{'Branch':{'$nin':old_branch}},{'Semester':{'$nin':old_semester}}],'Date':False})
                    original=course.copy()
                    course['Date']=Dates[iterator]
                    course['Time']=Type_thing.find_one()['Type']
                    new=course.copy()
                    course_table.find_one_and_replace(original,new)
                    old_branch.append(course['Branch'])
                    old_semester.append(course['Semester'])   
                iterator=(iterator+1)%len(Dates)
        else:
            iterator1=0
            iterator2=0
            while(course_table.find_one({'Time':False})):
                course=course_table.find_one({'Date':False})
                old=course.copy()
                old_branch=[]
                old_semester=[]
                course['Date']=Dates[iterator1]
                course['Time']=Times[iterator2]
                old_branch.append(course['Branch'])
                old_semester.append(course['Semester'])
                new=course.copy()
                course_table.find_one_and_replace(old,new)
                while(list(course_table.find({'$or':[{'Branch':{'$nin':old_branch}},{'Semester':{'$nin':old_semester}}],'Date':False}))):
                    course=course_table.find_one({'$or':[{'Branch':{'$nin':old_branch}},{'Semester':{'$nin':old_semester}}],'Date':False})
                    original=course.copy()
                    course['Date']=Dates[iterator1]
                    course['Time']=Times[iterator2]
                    new=course.copy()
                    course_table.find_one_and_replace(original,new)
                    old_branch.append(course['Branch'])
                    old_semester.append(course['Semester'])
                if iterator2==1:
                    iterator1=(iterator1+1)%len(Dates)
                    iterator2=0
                else:
                    iterator2=(iterator2+1)%len(Times)
            # for collection in course_table.find({},{'_id':False}).sort('Branch'):
            #     print(collection)
        course_name=None
        course_code=None
        for room in Room_table.find({'$or':[{'Capacity':{'$ne':60}},{'Capacity':{'$ne':40}}]}):
            usn=Room_student.find_one({'Room':room['Room Number'],'block':room['Block']})
            for course in course_table.find({'Branch':usn['Branch'],'Semester':usn['Semester']}):
                teacher=Teacher_table.find_one({'Number of Duties':{'$ne':0}})
                old=teacher.copy()
                if teacher['Number of Duties']>0:     
                    teacher['Number of Duties']=teacher['Number of Duties']-1
                    course_name=course['Course']
                    course_code=course['Course_code'] 
                    Time=course['Time']
                    Date=course['Date']
                    dictionary={'Name':teacher['Name'],'Subject':course_name,"Code":course_code,'Date':Date,'Time':Time,'Room':room["Room Number"],'Block':room['Block']}
                    Teacher_room_table.insert_one(dictionary)
                    new=teacher.copy()
                    Teacher_table.find_one_and_replace(old,new)
            usn1=Room_student.find_one({'Room':room['Room Number'],'block':room['Block'],'$or':[{'Branch':{'$ne':usn['Branch']}},{'Semester':{'$ne':usn['Semester']}}]})
            for course in course_table.find({'Branch':usn1['Branch'],'Semester':usn1['Semester']}):
                teacher=Teacher_table.find_one({'Number of Duties':{'$ne':0}})
                old=teacher.copy()
                if teacher['Number of Duties']>0:
                    teacher['Number of Duties']=teacher['Number of Duties']-1
                    course_name=course['Course']
                    course_code=course['Course_code'] 
                    Time=course['Time']
                    Date=course['Date']
                    dictionary={'Name':teacher['Name'],'Subject':course_name,"Code":course_code,'Date':Date,'Time':Time,'Room':room["Room Number"],'Block':room['Block']}
                    Teacher_room_table.insert_one(dictionary)
                    new=teacher.copy()
                    Teacher_table.find_one_and_replace(old,new)
class Output():
    def __init__(self):
        Output_sheet=openpyxl.load_workbook("Output.xlsx")
        for room in Room_table.find({'$or':[{'Capacity':{'$ne':60}},{'Capacity':{'$ne':40}}]}):
            sheets=Output_sheet.create_sheet()
            sheets.cell(1,1).value=room['Room Number']
            sheets.cell(1,2).value='Block'+room['Block']
            row=2
            for students in Room_student.find({'Room':room['Room Number'],'block':room['Block']}):
                student=student_table.find_one({'USN':students['student']})
                sheets.cell(row,1).value=student['Name']
                sheets.cell(row,2).value=student['USN']
                row=row+1
        sheet=Output_sheet.create_sheet()
        sheet.cell(1,1).value='Course'
        sheet.cell(1,2).value='Course Code'
        sheet.cell(1,3).value='Date'
        sheet.cell(1,4).value='Time'
        sheet.cell(1,5).value='Course Branch'
        sheet.cell(1,6).value='Semester'
        row1=2
        for course in course_table.find():
            sheet.cell(row1,1).value=course['Course']
            sheet.cell(row1,2).value=course['Course_code']
            sheet.cell(row1,3).value=course['Date']
            sheet.cell(row1,4).value=course['Time']
            sheet.cell(row1,5).value=course['Branch']
            sheet.cell(row1,6).value=course['Course']
            row1=row1+1
        last_sheet=Output_sheet.create_sheet()
        last_sheet.cell(1,1).value='Teacher Name'
        last_sheet.cell(1,2).value='Room Number'
        last_sheet.cell(1,3).value='Block'
        last_sheet.cell(1,4).value='Date of Examination'
        last_sheet.cell(1,5).value='Time of Examination'
        last_sheet.cell(1,6).value='Course'
        last_sheet.cell(1,7).value='Course code'
        row3=2
        for teacher in Teacher_room_table.find():
            last_sheet.cell(row3,1).value=teacher['Name']
            last_sheet.cell(row3,2).value=teacher['Room']
            last_sheet.cell(row3,3).value=teacher['Block']
            last_sheet.cell(row3,4).value=teacher['Date']
            last_sheet.cell(row3,5).value=teacher['Time']
            last_sheet.cell(row3,6).value=teacher['Subject'].replace('\n',' ')
            sheet.cell(row3,7).value=teacher['Code']
            row3=row3+1
        for date in Dates:
            date_sheet=Output_sheet.create_sheet()
            date_sheet.cell(1,1).value='Date'
            date_sheet.cell(1,2).value=date
            date_sheet.cell(2,1).value='Teacher Name'
            date_sheet.cell(2,2).value='Room Number'
            date_sheet.cell(2,3).value='Block'
            date_sheet.cell(2,4).value='Date of Examination'
            date_sheet.cell(2,5).value='Time of Examination'
            date_sheet.cell(2,6).value='Course'
            date_sheet.cell(2,7).value='Course code'
            row4=3
            for teacher in Teacher_room_table.find({'Date':date}):
                date_sheet.cell(row4,1).value=teacher['Name']
                date_sheet.cell(row4,2).value=teacher['Room']
                date_sheet.cell(row4,3).value=teacher['Block']
                date_sheet.cell(row4,4).value=teacher['Date']
                date_sheet.cell(row4,5).value=teacher['Time']
                date_sheet.cell(row4,6).value=teacher['Subject'].replace('\n',' ')
                date_sheet.cell(row4,7).value=teacher['Code']
                row4=row4+1
        Output_sheet.save('Output.xlsx')
        print('Done')        
Example=Database() 
Sample=Allocation()
Sample.Type_Dates()
Sample.function()
Over=Output()

# Sample.function()
